import openpyxl
from PIL import Image
import easyocr.model.vgg_model
import numpy as np
import re
import datetime
import time

# 找不到修改easyocr的源码
from openpyxl.styles import PatternFill

reader = easyocr.Reader(['ch_sim'], model_storage_directory="./", download_enabled=False, gpu=False, base_directory="./")

if __name__ == "__main__":

    # 匹配失败颜色
    Color = ['ffc7ce', '9c0006']
    # 设置填充颜色
    fills = PatternFill('solid', fgColor=Color[0])
    dates = [str, str, str, str, str, str, str]

    wb = openpyxl.load_workbook('./test2.xlsx')
    ws = wb[wb.sheetnames[0]]

    print("请输入匹配的日期格式  2022-04-12 :\n")
    thisTime = input()
    timeTuple = time.strptime(thisTime, "%Y-%m-%d")

    # 获取输入星期几
    tm_wday = timeTuple.tm_wday

    today = datetime.date.fromtimestamp(time.mktime(timeTuple))
    today = today - datetime.timedelta(days=tm_wday + 1)
    for i in range(7):
        today = today + datetime.timedelta(days=1)
        dates[i] = today.strftime("%Y-%m-%d")
    print("将匹配以下日期")
    print(dates)

    for image in ws._images:
        # 获取图片位置
        site = image.anchor._from
        # 获取图片坐标行
        row = site.row + 1
        # 转换可以识别的格式
        img = Image.open(image.ref).convert("RGB")
        img = np.array(img)

        sign = False

        print(ws.cell(row=row, column=3).value)
        # 读取图像
        result = reader.readtext(img)
        # 结果
        for i in result:

            mat = re.search(r"(\d{4}-\d{1,2}-\d{1,2})", i[1])
            if mat is not None:
                if mat.group() in dates:
                    print(mat.group() + " 匹配成功")
                    ws.cell(row=row, column=10).value = "匹配成功"

                    ws.cell(row=row, column=11).value = mat.group()
                    print("准确率" + str(i[2]))
                    ws.cell(row=row, column=12).value = "准确率" + str(i[2])

                    break


        # 序列
        ws.cell(row=row, column=10, value="匹配失败：图片时间不对").fill = fills

        print()
        print()
        print()

    for i in range(1, ws.max_row):
        if ws.cell(row=i, column=6).value == "是":
            if ws.cell(row=i, column=10).value is None:
                ws.cell(row=i, column=10, value="匹配失败：,没有图片").fill = fills

    wb.save("test2.xlsx")
